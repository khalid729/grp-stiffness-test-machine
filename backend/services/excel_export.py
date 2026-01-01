from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from typing import List
import logging

from db.models import Test

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export test data to Excel format"""

    def __init__(self):
        # Styles
        self.header_font = Font(bold=True, color='FFFFFF', size=11)
        self.header_fill = PatternFill(start_color='2C5282', end_color='2C5282', fill_type='solid')
        self.pass_fill = PatternFill(start_color='C6F6D5', end_color='C6F6D5', fill_type='solid')
        self.fail_fill = PatternFill(start_color='FED7D7', end_color='FED7D7', fill_type='solid')
        self.border = Border(
            left=Side(style='thin', color='CBD5E0'),
            right=Side(style='thin', color='CBD5E0'),
            top=Side(style='thin', color='CBD5E0'),
            bottom=Side(style='thin', color='CBD5E0')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')

    def export_tests(self, tests: List[Test]) -> bytes:
        """Export list of tests to Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"

        # Headers
        headers = [
            'ID', 'Date', 'Sample ID', 'Operator',
            'Diameter (mm)', 'Length (mm)', 'Deflection %',
            'Force@Target (kN)', 'Max Force (kN)', 'Ring Stiffness (kN/m²)',
            'SN Class', 'Result'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border

        # Data rows
        for row_num, test in enumerate(tests, 2):
            data = [
                test.id,
                test.test_date.strftime('%Y-%m-%d %H:%M') if test.test_date else '',
                test.sample_id or '',
                test.operator or '',
                test.pipe_diameter,
                test.pipe_length,
                test.deflection_percent,
                round(test.force_at_target, 2) if test.force_at_target else '',
                round(test.max_force, 2) if test.max_force else '',
                round(test.ring_stiffness, 0) if test.ring_stiffness else '',
                f"SN {test.sn_class}" if test.sn_class else '',
                'PASS' if test.passed else 'FAIL'
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.alignment = self.center_align
                cell.border = self.border

                # Color the result column
                if col == len(data):  # Result column
                    cell.fill = self.pass_fill if test.passed else self.fail_fill
                    cell.font = Font(bold=True)

        # Adjust column widths
        column_widths = [8, 18, 15, 15, 14, 14, 14, 18, 16, 20, 12, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add summary sheet
        self._add_summary_sheet(wb, tests)

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def _add_summary_sheet(self, wb: Workbook, tests: List[Test]):
        """Add summary statistics sheet"""
        ws = wb.create_sheet("Summary")

        # Calculate statistics
        total_tests = len(tests)
        passed_tests = sum(1 for t in tests if t.passed)
        failed_tests = total_tests - passed_tests
        pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0

        avg_stiffness = sum(t.ring_stiffness for t in tests if t.ring_stiffness) / max(1, sum(1 for t in tests if t.ring_stiffness))

        # SN class distribution
        sn_distribution = {}
        for test in tests:
            if test.sn_class:
                key = f"SN {test.sn_class}"
                sn_distribution[key] = sn_distribution.get(key, 0) + 1

        # Write summary
        summary_data = [
            ['Test Summary Report', ''],
            ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['', ''],
            ['Statistics', 'Value'],
            ['Total Tests', total_tests],
            ['Passed', passed_tests],
            ['Failed', failed_tests],
            ['Pass Rate', f"{pass_rate:.1f}%"],
            ['Average Ring Stiffness', f"{avg_stiffness:.0f} kN/m²"],
            ['', ''],
            ['SN Class Distribution', ''],
        ]

        for sn_class, count in sorted(sn_distribution.items()):
            summary_data.append([sn_class, count])

        for row_num, (label, value) in enumerate(summary_data, 1):
            ws.cell(row=row_num, column=1, value=label)
            ws.cell(row=row_num, column=2, value=value)

            if row_num == 1:
                ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
            elif label in ['Statistics', 'SN Class Distribution']:
                ws.cell(row=row_num, column=1).font = Font(bold=True)
                ws.cell(row=row_num, column=1).fill = self.header_fill
                ws.cell(row=row_num, column=1).font = Font(bold=True, color='FFFFFF')

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def export_test_with_data_points(self, test: Test) -> bytes:
        """Export single test with all data points"""
        wb = Workbook()

        # Test info sheet
        ws_info = wb.active
        ws_info.title = "Test Info"

        info_data = [
            ['Test Report', ''],
            ['', ''],
            ['Test ID', test.id],
            ['Date', test.test_date.strftime('%Y-%m-%d %H:%M') if test.test_date else ''],
            ['Sample ID', test.sample_id or ''],
            ['Operator', test.operator or ''],
            ['', ''],
            ['Parameters', ''],
            ['Pipe Diameter (mm)', test.pipe_diameter],
            ['Sample Length (mm)', test.pipe_length],
            ['Deflection Target (%)', test.deflection_percent],
            ['', ''],
            ['Results', ''],
            ['Force at Target (kN)', test.force_at_target],
            ['Max Force (kN)', test.max_force],
            ['Ring Stiffness (kN/m²)', test.ring_stiffness],
            ['SN Class', f"SN {test.sn_class}" if test.sn_class else ''],
            ['Result', 'PASS' if test.passed else 'FAIL'],
        ]

        for row_num, (label, value) in enumerate(info_data, 1):
            cell_label = ws_info.cell(row=row_num, column=1, value=label)
            cell_value = ws_info.cell(row=row_num, column=2, value=value)

            if label in ['Test Report', 'Parameters', 'Results']:
                cell_label.font = Font(bold=True, size=12)

        ws_info.column_dimensions['A'].width = 25
        ws_info.column_dimensions['B'].width = 20

        # Data points sheet
        if test.data_points:
            ws_data = wb.create_sheet("Data Points")

            headers = ['Time (s)', 'Force (kN)', 'Deflection (mm)', 'Position (mm)']
            for col, header in enumerate(headers, 1):
                cell = ws_data.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_align

            for row_num, dp in enumerate(sorted(test.data_points, key=lambda x: x.timestamp), 2):
                ws_data.cell(row=row_num, column=1, value=round(dp.timestamp, 3))
                ws_data.cell(row=row_num, column=2, value=round(dp.force, 3))
                ws_data.cell(row=row_num, column=3, value=round(dp.deflection, 3))
                ws_data.cell(row=row_num, column=4, value=round(dp.position, 3) if dp.position else '')

            for col in range(1, 5):
                ws_data.column_dimensions[get_column_letter(col)].width = 15

            ws_data.freeze_panes = 'A2'

        # Save
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
